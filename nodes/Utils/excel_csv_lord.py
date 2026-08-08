"""Read the first two columns of Excel/CSV files as prompt strings."""

import csv
import os
import posixpath
import zipfile
import xml.etree.ElementTree as ET
from typing import List, Tuple

try:
    import folder_paths
except ImportError:  # Allows syntax/import checks outside ComfyUI.
    folder_paths = None


SUPPORTED_EXTENSIONS = (".xlsx", ".xls", ".csv")


def _input_files() -> List[str]:
    if folder_paths is None:
        return []
    try:
        input_dir = folder_paths.get_input_directory()
        files = []
        for root, _, names in os.walk(input_dir):
            for name in names:
                if name.lower().endswith(SUPPORTED_EXTENSIONS):
                    files.append(os.path.relpath(os.path.join(root, name), input_dir))
        return sorted(files)
    except Exception:
        return []


def _resolve_path(file_name: str, file_path: str) -> str:
    if file_name and file_name.strip():
        if folder_paths is None:
            raise RuntimeError("ComfyUI folder_paths 不可用，请使用文件路径输入")
        input_dir = folder_paths.get_input_directory()
        candidate = os.path.join(input_dir, file_name.strip())
        if os.path.isfile(candidate):
            return candidate
        # The upload widget can submit only a basename while the file lives in a subfolder.
        wanted = os.path.basename(file_name.strip())
        for root, _, names in os.walk(input_dir):
            if wanted in names:
                return os.path.join(root, wanted)
        raise FileNotFoundError(f"在 ComfyUI/input 中找不到文件：{file_name}")

    if file_path and file_path.strip():
        candidate = file_path.strip()
        if not os.path.isabs(candidate) and folder_paths is not None:
            candidate = os.path.join(folder_paths.get_input_directory(), candidate)
        if os.path.isfile(candidate):
            return candidate
        raise FileNotFoundError(f"文件不存在：{file_path}")
    raise ValueError("请选择或输入一个 .xlsx、.xls 或 .csv 文件")


def _read_rows(path: str) -> List[Tuple[str, str]]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        last_error = None
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                with open(path, "r", encoding=encoding, newline="") as handle:
                    return [(row[0], row[1] if len(row) > 1 else "")
                            for row in csv.reader(handle) if row and any(cell.strip() for cell in row)]
            except UnicodeDecodeError as exc:
                last_error = exc
        raise RuntimeError(f"CSV 编码无法识别：{last_error}")

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            return _read_xlsx_stdlib(path)
            raise RuntimeError("读取 .xlsx 需要安装 openpyxl：pip install openpyxl") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            return [("" if row[0] is None else str(row[0]),
                     "" if len(row) < 2 or row[1] is None else str(row[1]))
                    for row in sheet.iter_rows(values_only=True)
                    if row and any(value is not None and str(value).strip() for value in row[:2])]
        finally:
            workbook.close()

    if ext == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("读取 .xls 需要安装 xlrd：pip install xlrd") from exc
        workbook = xlrd.open_workbook(path, on_demand=True)
        try:
            sheet = workbook.sheet_by_index(0)
            rows = []
            for index in range(sheet.nrows):
                rows.append((str(sheet.cell_value(index, 0) or ""),
                             str(sheet.cell_value(index, 1) if sheet.ncols > 1 else "")))
            return rows
        finally:
            workbook.release_resources()

    raise ValueError("仅支持 .xlsx、.xls、.csv 文件")


def _read_xlsx_stdlib(path: str) -> List[Tuple[str, str]]:
    main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    with zipfile.ZipFile(path) as archive:
        shared = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            shared = ["".join(node.text or "" for node in item.iter("{%s}t" % main)) for item in root.findall("{%s}si" % main)]
        root = ET.fromstring(archive.read("xl/workbook.xml"))
        sheet = root.find("{%s}sheets/{%s}sheet" % (main, main))
        rid = sheet.attrib.get("{%s}id" % rel) if sheet is not None else None
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        target = next((item.attrib.get("Target") for item in rels if item.attrib.get("Id") == rid), None)
        if not target:
            return []
        root = ET.fromstring(archive.read(posixpath.normpath(posixpath.join("xl", target))))
        rows = {}
        for cell in root.findall(".//{%s}c" % main):
            ref = cell.attrib.get("r", "")
            col = 0 if ref.startswith("A") else 1 if ref.startswith("B") else None
            if col is None:
                continue
            value = cell.find("{%s}v" % main)
            text = "" if value is None else (value.text or "")
            if cell.attrib.get("t") == "s" and text.isdigit() and int(text) < len(shared):
                text = shared[int(text)]
            row_num = "".join(ch for ch in ref if ch.isdigit())
            if row_num:
                rows.setdefault(int(row_num), ["", ""])[col] = text
        return [(values[0], values[1]) for _, values in sorted(rows.items()) if any(value.strip() for value in values)]


class LLExcelCSVLord:
    """LL-Excel-CSV-Lord: convert the first two columns into prompt lines."""

    @classmethod
    def INPUT_TYPES(cls):
        files = _input_files() or [""]
        return {
            "required": {
                "skip_header": ("BOOLEAN", {"default": False}),
                "separator": ("STRING", {"default": ",", "multiline": False}),
            },
            "optional": {
                "excel_csv_file": (files, {"image_upload": True, "editable": True}),
                "file_path": ("STRING", {"default": "", "multiline": False}),
            },
        }

    RETURN_TYPES = ("LIST", "STRING")
    RETURN_NAMES = ("prompt_list", "prompt_text")
    FUNCTION = "load_prompts"
    OUTPUT_NODE = True
    CATEGORY = "LLAI/工具"

    @classmethod
    def INPUT_LABELS(cls):
        return {
            "excel_csv_file": "Excel/CSV 文件",
            "file_path": "文件路径（可选）",
            "skip_header": "跳过首行表头",
            "separator": "单词与释义分隔符",
        }

    @classmethod
    def IS_CHANGED(cls, excel_csv_file="", file_path="", skip_header=False, separator=","):
        try:
            path = _resolve_path(excel_csv_file, file_path)
            return (os.path.getmtime(path), os.path.getsize(path), skip_header, separator)
        except Exception:
            return float("nan")

    def load_prompts(self, excel_csv_file="", file_path="", skip_header=False, separator=","):
        path = _resolve_path(excel_csv_file, file_path)
        rows = _read_rows(path)
        if skip_header and rows:
            rows = rows[1:]

        prompts = []
        for word, meaning in rows:
            word = str(word).strip()
            meaning = str(meaning).strip()
            if word or meaning:
                prompts.append(f"{word}{separator}{meaning}" if meaning else word)
        if not prompts:
            raise ValueError("文件中没有有效的前两列数据")
        text = "\n".join(prompts)
        return {"ui": {"llai_prompt_list": [prompts]}, "result": (prompts, text)}


NODE_CLASS_MAPPINGS = {"LL-Excel-CSV-Lord": LLExcelCSVLord}
NODE_DISPLAY_NAME_MAPPINGS = {"LL-Excel-CSV-Lord": "LL-Excel-CSV-Lord"}
