"""Upload an Excel workbook and expose its non-empty cells as prompts."""

import os

from .excel_csv_lord import _input_files, _resolve_path


MAX_PROMPTS = 1000
SUPPORTED_EXTENSIONS = (".xlsx", ".xls")


def _excel_files():
    return [name for name in _input_files() if name.lower().endswith(SUPPORTED_EXTENSIONS)] or [""]


def _read_all_cells(path):
    extension = os.path.splitext(path)[1].lower()
    if extension == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            from .excel_csv_lord import _read_xlsx_stdlib
            return [str(value).strip() for row in _read_xlsx_stdlib(path) for value in row
                    if value is not None and str(value).strip()]

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            values = []
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if value is not None and str(value).strip():
                        values.append(str(value).strip())
            return values
        finally:
            workbook.close()

    if extension == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("读取 .xls 需要安装 xlrd") from exc

        workbook = xlrd.open_workbook(path, on_demand=True)
        try:
            sheet = workbook.sheet_by_index(0)
            return [str(cell.value).strip() for row in sheet.get_rows() for cell in row
                    if cell.value is not None and str(cell.value).strip()]
        finally:
            workbook.release_resources()

    raise ValueError("仅支持 .xlsx 或 .xls 文件")


class LLReadExcel:
    """Read every non-empty cell from the first worksheet as one prompt."""

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_file": (_excel_files(), {
                    "image_upload": True,
                    "editable": True,
                    "tooltip": "拖放 .xlsx 或 .xls 文件到此处上传",
                }),
            },
        }

    @classmethod
    def INPUT_LABELS(cls):
        return {"excel_file": "load excel"}

    RETURN_TYPES = ("LIST", "STRING", "INT", "STRING")
    RETURN_NAMES = ("prompt_list", "prompt_text", "count", "prompt_list_text")
    FUNCTION = "read_excel"
    CATEGORY = "LLAI/工具"

    @classmethod
    def IS_CHANGED(cls, excel_file=""):
        try:
            path = _resolve_path(excel_file, "")
            return (os.path.getmtime(path), os.path.getsize(path))
        except Exception:
            return float("nan")

    def read_excel(self, excel_file=""):
        if not excel_file or not excel_file.strip():
            raise ValueError("请先上传 Excel 文件")
        if not excel_file.lower().endswith(SUPPORTED_EXTENSIONS):
            raise ValueError("仅支持 .xlsx 或 .xls 文件")

        path = _resolve_path(excel_file, "")
        prompts = _read_all_cells(path)
        if not prompts:
            raise ValueError("Excel 第一个工作表中没有有效内容")
        if len(prompts) > MAX_PROMPTS:
            raise ValueError(f"Excel 共读取到 {len(prompts)} 条内容，超过最多 {MAX_PROMPTS} 条的限制")

        prompt_text = "\n".join(prompts)
        return {
            "ui": {"llai_prompt_list": [prompts]},
            "result": (prompts, prompt_text, len(prompts), prompt_text),
        }


NODE_CLASS_MAPPINGS = {"LL-Read-Excel": LLReadExcel}
NODE_DISPLAY_NAME_MAPPINGS = {"LL-Read-Excel": "LL-Read-Excel"}
